import sys
import json

from collections import namedtuple

import openpyxl

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active

T = None
rows = []

def sanitize(s):
    return "".join(("", c)[c.isalnum()] for c in s)

for row in sheet:
    values = [c.value for c in row]
    if T is None:
        T = namedtuple("property", [sanitize(v) for v in values])
    else:
        rows.append(T(*values))

actions = []

for r in rows:
    name = r.name
    if r.newname is not None:
        name = r.newname
        actions.append(("rename", r.psetnameentity, r.name, r.newname))
    if r.newtypearguments is not None:
        actions.append(("change", r.psetnameentity, r.name, r.originaltypearguments, r.newtypearguments))

wb = openpyxl.Workbook()
sheet = wb.active
for i, a in enumerate(sorted(actions), start=1):
    for j, c in enumerate(a, start=1):
        sheet.cell(row=i, column=j).value = c
        
wb.save("actions.xlsx")

json.dump([a for a in actions if a[0] == 'rename'], open("actions.json", "w"),indent=1)

# breakpoint()
